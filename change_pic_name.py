import os
import openpyxl
import requests


class ExcelWork:
    def __init__(self):
        self.list_tables = []
        self.old_link_dict = {}
        self.new_link_dict = {}

    @staticmethod
    def find_col_num(col_name, work_sheet):
        col_num = -1
        for col in work_sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            col_num += 1
            if col_name in col[0]:
                break
        return col_num

    @staticmethod
    def photo_saver(url_, name, path=None):
        response = requests.get(url_)
        with open(f'{path}', 'wb') as f:
            f.write(response.content)

    def find_file(self, dir_path_):
        files = os.listdir(dir_path_)
        self.list_tables = [file for file in files if ".xlsx"]

        return self.list_tables

    def get_links(self, table_path_):
        wb = openpyxl.load_workbook(table_path_)
        ws = wb.active

        link_col_num = self.find_col_num('фото', ws)
        code_col_num = self.find_col_num('Код товара', ws)

        row_num = 1
        for row in ws.iter_rows(min_col=link_col_num + 1, max_col=link_col_num + 1, min_row=2, values_only=True):
            row_num += 1
            if row[0]:
                product_links = [link.strip() for link in row[0].split(',')]
            else:
                product_links = 0

            product_code = ws.cell(row_num, code_col_num + 1).value
            self.old_link_dict[product_code] = product_links
        return self.old_link_dict

    def links_rename(self, table_path_, photo_dir_path=None):
        if self.old_link_dict:
            wb = openpyxl.load_workbook(table_path_)
            ws = wb.active

            link_col_num = self.find_col_num('фото', ws)

            row = 2
            for key, values in self.old_link_dict.items():
                if not key:
                    break
                if not values:
                    continue
                amount_link = len(values)

                link_list = [f"{key}-{num}" for num in range(1, amount_link + 1)]
                num = 1
                for url in values:
                    self.photo_saver(url, f"{key}-{num}", photo_dir_path)
                    num += 1

                self.new_link_dict[key] = link_list
                link_list = [f"https://polezniemelochi.ru/wp-content/uploads/photo/{link}" for link in link_list]
                ws.cell(row=row, column=link_col_num + 1).value = ', '.join(link_list)
                row += 1
            wb.save(table_path_)


if __name__ == "__main__":
    table_dir_path = 'таблицы'
    photo_dir_path_ = 'hosting/polezniemelochi.ru/wp-content/uploads/photo'
    ew_ = ExcelWork()
    tables_list = ew_.find_file(table_dir_path)
    for table_path in tables_list:
        try:
            link_dict_ = ew_.get_links(f"{table_dir_path}/{table_path}")
            ew_.links_rename(f"{table_dir_path}/{table_path}", photo_dir_path=photo_dir_path_)
        except TypeError as E:
            print(f"An error: ####{E}#### \n occurred while processing table {table_path}.\nThe iteration was skipped!")


